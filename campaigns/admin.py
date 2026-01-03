from django.contrib import admin, messages
from django.shortcuts import redirect, render
from django.urls import path

from openpyxl import load_workbook

from campaigns.models import Campaign, CampaignRecipient, EmailEvent, Recipient, RecipientTag


@admin.register(Campaign)
class CampaignAdmin(admin.ModelAdmin):
    list_display = ("name", "start_at", "end_at", "throttle_per_minute", "created_by")
    search_fields = ("name",)
    list_filter = ("start_at", "end_at")
    filter_horizontal = ("recipient_tags",)


@admin.register(Recipient)
class RecipientAdmin(admin.ModelAdmin):
    list_display = ("email", "full_name", "role", "area", "department", "created_at")
    search_fields = ("email", "full_name", "role", "area", "department")
    filter_horizontal = ("tags",)
    change_list_template = "admin/campaigns/recipient/change_list.html"

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path(
                "import-users/",
                self.admin_site.admin_view(self.import_users),
                name="campaigns_recipient_import",
            ),
        ]
        return custom_urls + urls

    def import_users(self, request):
        if request.method == "POST" and request.FILES.get("excel_file"):
            workbook = load_workbook(request.FILES["excel_file"])
            sheet = workbook.active
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                self.message_user(request, "El archivo no contiene filas.", level=messages.ERROR)
                return redirect("..")

            header = [str(value or "").strip().lower() for value in rows[0]]
            column_map = {name: index for index, name in enumerate(header)}
            required = ["email", "nombres", "apellido paterno", "apellido materno", "rol", "area"]
            missing = [name for name in required if name not in column_map]
            if missing:
                self.message_user(
                    request,
                    f"Faltan columnas requeridas: {', '.join(missing)}.",
                    level=messages.ERROR,
                )
                return redirect("..")

            created_count = 0
            updated_count = 0
            for row in rows[1:]:
                email = str(row[column_map["email"]] or "").strip()
                if not email:
                    continue
                first_name = str(row[column_map["nombres"]] or "").strip()
                last_name_paternal = str(row[column_map["apellido paterno"]] or "").strip()
                last_name_maternal = str(row[column_map["apellido materno"]] or "").strip()
                role = str(row[column_map["rol"]] or "").strip()
                area = str(row[column_map["area"]] or "").strip()
                full_name = " ".join(
                    part for part in [first_name, last_name_paternal, last_name_maternal] if part
                ).strip()
                defaults = {
                    "first_name": first_name,
                    "last_name_paternal": last_name_paternal,
                    "last_name_maternal": last_name_maternal,
                    "role": role,
                    "area": area,
                    "department": area,
                    "full_name": full_name,
                }
                recipient, created = Recipient.objects.update_or_create(email=email, defaults=defaults)
                if created:
                    created_count += 1
                else:
                    updated_count += 1

            self.message_user(
                request,
                f"Importación completada. Nuevos: {created_count}, actualizados: {updated_count}.",
                level=messages.SUCCESS,
            )
            return redirect("..")

        return render(request, "admin/campaigns/recipient/import.html")


@admin.register(RecipientTag)
class RecipientTagAdmin(admin.ModelAdmin):
    list_display = ("name", "created_at")
    search_fields = ("name",)


@admin.register(CampaignRecipient)
class CampaignRecipientAdmin(admin.ModelAdmin):
    list_display = (
        "campaign",
        "recipient",
        "status",
        "sent_at",
        "opened_at",
        "clicked_at",
        "click_count",
        "landing_viewed_at",
        "landing_view_count",
    )
    list_filter = ("status", "opened_at", "clicked_at")
    search_fields = ("campaign__name", "recipient__email")

    def save_model(self, request, obj, form, change) -> None:
        obj.full_clean()
        super().save_model(request, obj, form, change)


@admin.register(EmailEvent)
class EmailEventAdmin(admin.ModelAdmin):
    list_display = (
        "event_type",
        "recipient",
        "device_type",
        "os_family",
        "browser_family",
        "email_client_hint",
        "created_at",
    )
    list_filter = ("event_type", "created_at")
    search_fields = ("recipient__recipient__email", "recipient__campaign__name")
